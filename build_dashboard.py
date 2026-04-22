import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, RadarChart, Reference
from openpyxl.chart.series import DataPoint
import warnings
warnings.filterwarnings('ignore')

print("=" * 60)
print("NBA OPTIMIZER — EXCEL DASHBOARD BUILDER")
print("=" * 60)

# ============================================================
# LOAD ALL DATA
# ============================================================

df_all    = pd.read_csv('nba_clustered.csv')
df_scored = pd.read_csv('nba_scored_complete.csv')
df_2324   = df_all[df_all['SEASON'] == '2023-24'].copy()

roster_A = pd.read_csv('optimized_roster_A.csv')
roster_B = pd.read_csv('optimized_roster_B.csv')
roster_C = pd.read_csv('optimized_roster_C.csv')

print(f"Loaded {len(df_all)} total player-seasons")
print(f"Loaded 3 optimized rosters")

# ============================================================
# STYLE HELPERS
# ============================================================

NAVY       = "1B2A4A"
GOLD       = "C9A84C"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
DARK_GRAY  = "2C2C2C"
LIGHT_GRAY = "F2F2F2"
GREEN      = "27AE60"
RED        = "E74C3C"
ORANGE     = "F39C12"

def header_cell(ws, row, col, value, bg=NAVY, fg=WHITE, bold=True, size=11):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    return cell

def data_cell(ws, row, col, value, bg=WHITE, bold=False, align='center', fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    if fmt:
        cell.number_format = fmt
    return cell

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def freeze_and_filter(ws, row, col):
    ws.freeze_panes = ws.cell(row=row, column=col)

wb = Workbook()

# ============================================================
# TAB 1 — COVER PAGE
# ============================================================

ws_cover = wb.active
ws_cover.title = "Cover"
ws_cover.sheet_view.showGridLines = False

ws_cover.row_dimensions[1].height = 20
ws_cover.row_dimensions[2].height = 60
ws_cover.row_dimensions[3].height = 40
ws_cover.row_dimensions[4].height = 30
ws_cover.row_dimensions[5].height = 25
ws_cover.row_dimensions[6].height = 25
ws_cover.row_dimensions[7].height = 25
ws_cover.row_dimensions[8].height = 25
ws_cover.row_dimensions[9].height = 25
ws_cover.row_dimensions[10].height = 25

ws_cover.merge_cells('A2:H2')
title = ws_cover['A2']
title.value = "NBA ROSTER OPTIMIZATION ENGINE"
title.font = Font(bold=True, size=28, color=GOLD)
title.fill = PatternFill("solid", fgColor=NAVY)
title.alignment = Alignment(horizontal='center', vertical='center')

ws_cover.merge_cells('A3:H3')
sub = ws_cover['A3']
sub.value = "Building the Statistically Optimal Team Under a Salary Cap"
sub.font = Font(bold=True, size=16, color=WHITE)
sub.fill = PatternFill("solid", fgColor=NAVY)
sub.alignment = Alignment(horizontal='center', vertical='center')

ws_cover.merge_cells('A5:H5')
ws_cover['A5'].value = "Project Overview"
ws_cover['A5'].font = Font(bold=True, size=13, color=NAVY)
ws_cover['A5'].alignment = Alignment(horizontal='left')

rows_overview = [
    ("Data",        "399 NBA players × 3 seasons (2021-22 → 2023-24) = 1,208 player-seasons"),
    ("Method",      "Three-framework Pearson + Spearman correlation → PCA composite score → K-Means clustering → MILP optimization"),
    ("Key Stats",   "9 stats validated across 3 correlation frameworks: OFF/DEF rating (team-adjusted), PIE, BLK, STL, DREB%, AST%, TS%, FG3%"),
    ("Archetypes",  "7 player archetypes discovered via K-Means: Elite Two-Way Big, Rim Protector, Perimeter Playmaker, 3-and-D Wing, Stretch Big, Defensive Specialist, Bench Player"),
    ("Output",      "3 optimized 15-man rosters under $136M (hard cap), $165M (luxury tax), $90M (budget) salary constraints"),
]

for i, (label, desc) in enumerate(rows_overview):
    r = 6 + i
    ws_cover.merge_cells(f'A{r}:B{r}')
    ws_cover.merge_cells(f'C{r}:H{r}')
    ws_cover[f'A{r}'].value = label
    ws_cover[f'A{r}'].font = Font(bold=True, size=10, color=NAVY)
    ws_cover[f'C{r}'].value = desc
    ws_cover[f'C{r}'].font = Font(size=10)
    ws_cover[f'C{r}'].alignment = Alignment(wrap_text=True)

for col in range(1, 9):
    set_col_width(ws_cover, col, 18)

print("Tab 1 (Cover) done")

# ============================================================
# TAB 2 — MASTER PLAYER DATABASE (2023-24)
# ============================================================

ws_db = wb.create_sheet("Player Database")
ws_db.sheet_view.showGridLines = False

cols_db = ['PLAYER_NAME', 'TEAM_ABBREVIATION', 'ARCHETYPE',
           'COMPOSITE_SCORE_NORM', 'SALARY', 'VORPD',
           'OFF_RATING_ADJUSTED', 'DEF_RATING_ADJUSTED',
           'TS_PCT', 'BLK', 'STL', 'AST_PCT_ADJUSTED',
           'DREB_PCT_ADJUSTED', 'FG3_PCT', 'PIE_ADJUSTED']

col_labels = ['Player', 'Team', 'Archetype', 'Score (0-100)',
              'Salary', 'VORPD', 'OFF Rtg Adj',
              'DEF Rtg Adj', 'TS%', 'BLK', 'STL',
              'AST% Adj', 'DREB% Adj', 'FG3%', 'PIE Adj']

# Title row
ws_db.merge_cells('A1:O1')
header_cell(ws_db, 1, 1, "NBA Player Database — 2023-24 Season (399 Players)", bg=NAVY, size=13)
ws_db.row_dimensions[1].height = 30

# Column headers
for c, label in enumerate(col_labels, 1):
    header_cell(ws_db, 2, c, label, bg=DARK_GRAY)
ws_db.row_dimensions[2].height = 35

# Data rows
df_db = df_2324[cols_db].sort_values('COMPOSITE_SCORE_NORM', ascending=False).reset_index(drop=True)

archetype_colors = {
    'Elite Two-Way Big':    'D5E8D4',
    'Rim Protector':        'DAE8FC',
    'Perimeter Playmaker':  'FFE6CC',
    'Stretch Big':          'E1D5E7',
    '3-and-D Wing':         'FFF2CC',
    'Defensive Specialist': 'F8CECC',
    'Bench / Fringe Player':'F5F5F5',
}

for r_idx, row in df_db.iterrows():
    row_num = r_idx + 3
    bg = archetype_colors.get(row['ARCHETYPE'], WHITE)
    alt_bg = WHITE if r_idx % 2 == 0 else LIGHT_GRAY

    for c_idx, col in enumerate(cols_db, 1):
        val = row[col]
        if col == 'SALARY':
            data_cell(ws_db, row_num, c_idx, val, bg=bg, fmt='$#,##0')
        elif col in ['COMPOSITE_SCORE_NORM', 'VORPD']:
            data_cell(ws_db, row_num, c_idx, round(float(val), 2) if pd.notna(val) else '', bg=bg)
        elif col in ['TS_PCT', 'FG3_PCT']:
            data_cell(ws_db, row_num, c_idx, round(float(val), 3) if pd.notna(val) else '', bg=bg, fmt='0.000')
        elif col in ['OFF_RATING_ADJUSTED', 'DEF_RATING_ADJUSTED',
                     'AST_PCT_ADJUSTED', 'DREB_PCT_ADJUSTED', 'PIE_ADJUSTED']:
            data_cell(ws_db, row_num, c_idx, round(float(val), 2) if pd.notna(val) else '', bg=bg, fmt='0.00')
        elif col in ['BLK', 'STL']:
            data_cell(ws_db, row_num, c_idx, round(float(val), 2) if pd.notna(val) else '', bg=bg, fmt='0.00')
        else:
            data_cell(ws_db, row_num, c_idx, val, bg=bg, align='left' if c_idx <= 3 else 'center')

widths_db = [22, 7, 22, 13, 14, 8, 11, 11, 7, 6, 6, 9, 10, 7, 9]
for c, w in enumerate(widths_db, 1):
    set_col_width(ws_db, c, w)

freeze_and_filter(ws_db, 3, 1)
ws_db.auto_filter.ref = f"A2:O{len(df_db)+2}"
print("Tab 2 (Player Database) done")

# ============================================================
# TAB 3 — OPTIMIZED ROSTERS (ALL THREE SCENARIOS)
# ============================================================

ws_ros = wb.create_sheet("Optimized Rosters")
ws_ros.sheet_view.showGridLines = False

scenario_data = [
    ("SCENARIO A — Hard Cap ($136M)",   roster_A, 1,  136_021_000, NAVY),
    ("SCENARIO B — Luxury Tax ($165M)", roster_B, 21, 165_294_000, "1A5276"),
    ("SCENARIO C — Budget Team ($90M)", roster_C, 41, 90_000_000,  "1E8449"),
]

for title_text, roster, start_row, cap, color in scenario_data:
    ws_ros.merge_cells(f'A{start_row}:G{start_row}')
    header_cell(ws_ros, start_row, 1, title_text, bg=color, size=12)
    ws_ros.row_dimensions[start_row].height = 28

    headers = ['Player', 'Team', 'Archetype', 'Score', 'Salary', 'VORPD', 'Cap %']
    for c, h in enumerate(headers, 1):
        header_cell(ws_ros, start_row + 1, c, h, bg=DARK_GRAY)

    total_salary = roster['SALARY'].sum()
    total_score = roster['COMPOSITE_SCORE_NORM'].sum()

    for i, (_, row) in enumerate(roster.sort_values('COMPOSITE_SCORE_NORM', ascending=False).iterrows()):
        r = start_row + 2 + i
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        cap_pct = row['SALARY'] / cap * 100

        data_cell(ws_ros, r, 1, row['PLAYER_NAME'], bg=bg, align='left')
        data_cell(ws_ros, r, 2, row['TEAM_ABBREVIATION'], bg=bg)
        data_cell(ws_ros, r, 3, row['ARCHETYPE'], bg=bg, align='left')
        data_cell(ws_ros, r, 4, round(row['COMPOSITE_SCORE_NORM'], 2), bg=bg)
        data_cell(ws_ros, r, 5, row['SALARY'], bg=bg, fmt='$#,##0')
        data_cell(ws_ros, r, 6, round(row['VORPD'], 2), bg=bg)
        data_cell(ws_ros, r, 7, round(cap_pct, 1), bg=bg, fmt='0.0"%"')

    # Summary row
    sum_row = start_row + 2 + len(roster)
    ws_ros.merge_cells(f'A{sum_row}:C{sum_row}')
    header_cell(ws_ros, sum_row, 1, "TOTALS", bg=color, fg=WHITE)
    data_cell(ws_ros, sum_row, 4, round(total_score, 2), bg=LIGHT_BLUE, bold=True)
    data_cell(ws_ros, sum_row, 5, total_salary, bg=LIGHT_BLUE, bold=True, fmt='$#,##0')
    cap_used = total_salary / cap * 100
    data_cell(ws_ros, sum_row, 7, round(cap_used, 1), bg=LIGHT_BLUE, bold=True, fmt='0.0"%"')

widths_ros = [24, 7, 24, 9, 16, 8, 8]
for c, w in enumerate(widths_ros, 1):
    set_col_width(ws_ros, c, w)

print("Tab 3 (Optimized Rosters) done")

# ============================================================
# TAB 4 — SCENARIO COMPARISON
# ============================================================

ws_cmp = wb.create_sheet("Scenario Comparison")
ws_cmp.sheet_view.showGridLines = False

ws_cmp.merge_cells('A1:E1')
header_cell(ws_cmp, 1, 1, "Scenario Comparison — What Does Extra Cap Space Buy?", bg=NAVY, size=13)
ws_cmp.row_dimensions[1].height = 30

headers_cmp = ['Metric', 'A — Hard Cap ($136M)', 'B — Luxury Tax ($165M)', 'C — Budget ($90M)', 'A vs C Difference']
for c, h in enumerate(headers_cmp, 1):
    header_cell(ws_cmp, 2, c, h, bg=DARK_GRAY)

scenarios_cmp = {
    'A': (roster_A, 136_021_000),
    'B': (roster_B, 165_294_000),
    'C': (roster_C, 90_000_000),
}

def get_stats(roster, cap):
    return {
        'total_score': roster['COMPOSITE_SCORE_NORM'].sum(),
        'avg_score':   roster['COMPOSITE_SCORE_NORM'].mean(),
        'max_score':   roster['COMPOSITE_SCORE_NORM'].max(),
        'total_sal':   roster['SALARY'].sum(),
        'cap_used':    roster['SALARY'].sum() / cap * 100,
        'avg_vorpd':   roster['VORPD'].mean(),
        'n_stars':     (roster['COMPOSITE_SCORE_NORM'] >= 70).sum(),
    }

sa = get_stats(*scenarios_cmp['A'])
sb = get_stats(*scenarios_cmp['B'])
sc = get_stats(*scenarios_cmp['C'])

metrics = [
    ('Total Composite Score',    sa['total_score'],  sb['total_score'],  sc['total_score'],  '$#,##0.00'),
    ('Average Player Score',     sa['avg_score'],    sb['avg_score'],    sc['avg_score'],    '0.00'),
    ('Best Player Score',        sa['max_score'],    sb['max_score'],    sc['max_score'],    '0.00'),
    ('Total Salary Used',        sa['total_sal'],    sb['total_sal'],    sc['total_sal'],    '$#,##0'),
    ('Cap % Used',               sa['cap_used'],     sb['cap_used'],     sc['cap_used'],     '0.0"%"'),
    ('Avg VORPD',                sa['avg_vorpd'],    sb['avg_vorpd'],    sc['avg_vorpd'],    '0.00'),
    ('Players Scoring 70+',      sa['n_stars'],      sb['n_stars'],      sc['n_stars'],      '0'),
]

for i, (label, val_a, val_b, val_c, fmt) in enumerate(metrics):
    r = i + 3
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    data_cell(ws_cmp, r, 1, label, bg=bg, bold=True, align='left')
    data_cell(ws_cmp, r, 2, round(val_a, 2) if isinstance(val_a, float) else val_a, bg=bg, fmt=fmt)
    data_cell(ws_cmp, r, 3, round(val_b, 2) if isinstance(val_b, float) else val_b, bg=bg, fmt=fmt)
    data_cell(ws_cmp, r, 4, round(val_c, 2) if isinstance(val_c, float) else val_c, bg=bg, fmt=fmt)
    diff = val_a - val_c if isinstance(val_a, (int, float)) else ''
    diff_cell = data_cell(ws_cmp, r, 5, round(diff, 2) if isinstance(diff, float) else diff, bg=bg)
    if isinstance(diff, (int, float)) and diff > 0:
        diff_cell.font = Font(color=GREEN, bold=True, size=10)

for c, w in enumerate([28, 22, 22, 22, 18], 1):
    set_col_width(ws_cmp, c, w)

print("Tab 4 (Scenario Comparison) done")

# ============================================================
# TAB 5 — ARCHETYPE SUMMARY
# ============================================================

ws_arch = wb.create_sheet("Archetypes")
ws_arch.sheet_view.showGridLines = False

ws_arch.merge_cells('A1:I1')
header_cell(ws_arch, 1, 1, "Player Archetypes — K-Means Clustering Results (7 Archetypes)", bg=NAVY, size=13)
ws_arch.row_dimensions[1].height = 30

arch_headers = ['Archetype', 'Players\n(3 seasons)', 'Players\n(2023-24)',
                'Avg Score', 'Avg Salary', 'Avg VORPD',
                'Avg OFF Rtg', 'Avg DEF Rtg', 'Defining Traits']
for c, h in enumerate(arch_headers, 1):
    header_cell(ws_arch, 2, c, h, bg=DARK_GRAY)
ws_arch.row_dimensions[2].height = 35

arch_traits = {
    'Elite Two-Way Big':    'High PIE, BLK, TS% — dominant on both ends',
    'Rim Protector':        'High BLK, DREB% — pure interior defense',
    'Perimeter Playmaker':  'High AST%, OFF Rating — ball-handling, creation',
    'Stretch Big':          'High FG3%, moderate BLK — versatile shooting big',
    '3-and-D Wing':         'High FG3% + STL — floor spacing + perimeter D',
    'Defensive Specialist': 'Highest DEF Rating — lockdown defensive players',
    'Bench / Fringe Player':'Below-average across stats — end-of-bench role',
}

archetypes_order = list(arch_traits.keys())
for i, arch in enumerate(archetypes_order):
    r = i + 3
    bg = list(archetype_colors.values())[i]

    players_all  = df_all[df_all['ARCHETYPE'] == arch]
    players_2324 = df_2324[df_2324['ARCHETYPE'] == arch]

    data_cell(ws_arch, r, 1, arch, bg=bg, bold=True, align='left')
    data_cell(ws_arch, r, 2, len(players_all), bg=bg)
    data_cell(ws_arch, r, 3, len(players_2324), bg=bg)
    data_cell(ws_arch, r, 4, round(players_2324['COMPOSITE_SCORE_NORM'].mean(), 1), bg=bg)
    data_cell(ws_arch, r, 5, int(players_2324['SALARY'].mean()), bg=bg, fmt='$#,##0')
    data_cell(ws_arch, r, 6, round(players_2324['VORPD'].mean(), 2), bg=bg)
    if 'OFF_RATING_ADJUSTED' in players_2324.columns:
        data_cell(ws_arch, r, 7, round(players_2324['OFF_RATING_ADJUSTED'].mean(), 2), bg=bg, fmt='0.00')
        data_cell(ws_arch, r, 8, round(players_2324['DEF_RATING_ADJUSTED'].mean(), 2), bg=bg, fmt='0.00')
    data_cell(ws_arch, r, 9, arch_traits[arch], bg=bg, align='left')

for c, w in enumerate([24, 10, 10, 10, 14, 10, 11, 11, 45], 1):
    set_col_width(ws_arch, c, w)

print("Tab 5 (Archetypes) done")

# ============================================================
# TAB 6 — METHODOLOGY
# ============================================================

ws_meth = wb.create_sheet("Methodology")
ws_meth.sheet_view.showGridLines = False

ws_meth.merge_cells('A1:F1')
header_cell(ws_meth, 1, 1, "Methodology — NBA Roster Optimization Engine", bg=NAVY, size=14)
ws_meth.row_dimensions[1].height = 35

methodology_text = [
    ("DATA COLLECTION", ""),
    ("Source", "NBA Stats API (nba_api Python package) — official NBA.com backend"),
    ("Seasons", "2021-22, 2022-23, 2023-24 (3 seasons × ~400 players = 1,208 player-seasons)"),
    ("Filter", "Minimum 20 games played and 10 minutes per game to exclude noise"),
    ("Salary", "Basketball Reference contracts scaled to historical seasons via cap-ratio method"),
    ("Recency Weights", "2021-22: 20%, 2022-23: 35%, 2023-24: 45% — recent seasons weighted more"),
    ("", ""),
    ("FEATURE ENGINEERING", ""),
    ("Team Adjustment", "All on-court ratings (OFF, DEF, PIE, AST%, DREB%) adjusted by subtracting team average to isolate individual contribution from teammate quality"),
    ("DEF Rating Flip", "DEF_RATING multiplied by -1 so higher always means better (lower DEF rating = better defense)"),
    ("TS% Filter", "Players with fewer than 100 total field goal attempts receive league average TS% to prevent small-sample distortion"),
    ("", ""),
    ("CORRELATION ANALYSIS (3 FRAMEWORKS)", ""),
    ("Framework 1", "Pearson correlation of team-adjusted stats vs W_PCT — identifies stats that predict team winning"),
    ("Framework 2", "Pearson correlation of individual counting stats vs W_PCT — identifies individual skills tied to winning"),
    ("Framework 3", "Pearson correlation of stats vs PLUS_MINUS_ADJUSTED — captures invisible contributions not in traditional stats"),
    ("Spearman Validation", "All candidate stats re-validated using Spearman rank correlation to catch non-linear relationships Pearson misses"),
    ("Selection Rule", "Stat included if p < 0.05 in at least one framework AND validated by Spearman. Avoids multiple comparisons inflation."),
    ("", ""),
    ("PCA COMPOSITE SCORE", ""),
    ("Standardization", "All 9 final stats normalized to mean=0, std=1 via StandardScaler before PCA — prevents larger-scale stats from dominating"),
    ("PCA Method", "Principal Component Analysis on 9 validated stats. First component weights derived entirely from data — no manual weight assignment"),
    ("Score Normalization", "Raw PCA scores rescaled to 0-100 for interpretability"),
    ("VORPD", "Value Over Replacement Per Dollar = (Score - 10th percentile score) / (Salary in $M). Measures salary cap efficiency."),
    ("", ""),
    ("K-MEANS CLUSTERING", ""),
    ("Method", "K-Means clustering (K=7) on all 9 standardized stats across 1,208 player-seasons"),
    ("K Selection", "Elbow method used to validate K=7 as the point of diminishing returns in inertia reduction"),
    ("Archetypes", "7 clusters manually labeled based on average stat profiles and basketball identity of cluster members"),
    ("Key Finding", "Model independently separated 'Rim Protectors' (low FG3%) from 'Stretch Bigs' (high FG3%) — validating that positionless clustering captures modern NBA reality"),
    ("", ""),
    ("MILP OPTIMIZATION", ""),
    ("Model Type", "Mixed Integer Linear Programming (MILP) via PuLP/CBC solver"),
    ("Objective", "Maximize sum of COMPOSITE_SCORE_NORM for selected 15 players"),
    ("Hard Constraints", "Salary cap, roster size (15), archetype minimums (see below), max 2 players per real NBA team"),
    ("Archetype Mins", "Perimeter Playmaker ≥ 2, 3-and-D Wing ≥ 2, all others ≥ 1. Big men capped at 7 total."),
    ("Scenarios", "A: $136M hard cap | B: $165M luxury tax | C: $90M budget"),
    ("", ""),
    ("LIMITATIONS", ""),
    ("Salary Approximation", "Historical salaries estimated via cap-ratio scaling from current contracts. ~384 player-seasons used imputed league minimum due to name-matching failures. Notably Jimmy Butler's salary was mismatched."),
    ("Big Man Bias", "PCA Component 1 is dominated by interior stats (PIE, BLK, DREB). Guards are undervalued in the composite score. Archetype constraints partially mitigate this."),
    ("Team Context Limitation", "PLUS_MINUS_ADJUSTED still contains lineup contamination bias that team-averaging cannot fully remove"),
    ("Single Season Optimization", "Optimizer uses 2023-24 data only. Multi-season player aging curves not modeled."),
]

for i, (label, text) in enumerate(methodology_text):
    r = i + 2
    if text == "" and label == "":
        ws_meth.row_dimensions[r].height = 8
        continue
    if text == "":
        ws_meth.merge_cells(f'A{r}:F{r}')
        cell = ws_meth[f'A{r}']
        cell.value = label
        cell.font = Font(bold=True, size=11, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK_GRAY)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws_meth.row_dimensions[r].height = 22
    else:
        ws_meth[f'A{r}'].value = label
        ws_meth[f'A{r}'].font = Font(bold=True, size=10, color=NAVY)
        ws_meth[f'A{r}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws_meth.merge_cells(f'B{r}:F{r}')
        ws_meth[f'B{r}'].value = text
        ws_meth[f'B{r}'].font = Font(size=10)
        ws_meth[f'B{r}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        ws_meth[f'A{r}'].fill = PatternFill("solid", fgColor=bg)
        ws_meth[f'B{r}'].fill = PatternFill("solid", fgColor=bg)
        ws_meth.row_dimensions[r].height = 30

set_col_width(ws_meth, 1, 28)
for c in range(2, 7):
    set_col_width(ws_meth, c, 20)

print("Tab 6 (Methodology) done")

# ============================================================
# SAVE WORKBOOK
# ============================================================

output_file = 'NBA_Optimizer_Dashboard.xlsx'
wb.save(output_file)
print(f"\n{'=' * 60}")
print(f"DASHBOARD SAVED: {output_file}")
print(f"{'=' * 60}")
print("\nTabs created:")
print("  1. Cover          — Project overview")
print("  2. Player Database — All 399 players, scored and color-coded by archetype")
print("  3. Optimized Rosters — All 3 salary cap scenarios")
print("  4. Scenario Comparison — Side-by-side metrics")
print("  5. Archetypes     — K-Means cluster summary")
print("  6. Methodology    — Full statistical methodology writeup")
print("\nOpen NBA_Optimizer_Dashboard.xlsx to review the final output.")
