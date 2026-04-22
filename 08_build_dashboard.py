import sys
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

print("=" * 60)
print("NBA OPTIMIZER - EXCEL DASHBOARD BUILDER")
print("=" * 60)

# ============================================================
# LOAD ALL DATA
# ============================================================

df_all  = pd.read_csv('nba_clustered.csv')
df_2324 = df_all[df_all['SEASON'] == '2023-24'].copy()

roster_files = {
    'A': ('A - Hard Cap ($136M)',        136_021_000, 'NAVY'),
    'B': ('B - Luxury Tax ($165M)',      165_294_000, 'BLUE'),
    'C': ('C - Budget Team ($90M)',       90_000_000, 'GREEN'),
    'D': ('D - Rebuild Mode ($136M)',    136_021_000, 'PURPLE'),
    'E': ('E - Win-Now ($165M)',         165_294_000, 'TEAL'),
    'F': ('F - Defensive Identity ($136M)', 136_021_000, 'RED'),
    'G': ('G - Offensive Identity ($136M)', 136_021_000, 'ORANGE'),
    'H': ('H - Three-Point Era ($136M)', 136_021_000, 'CYAN'),
    'I': ('I - Small Ball ($136M)',      136_021_000, 'BROWN'),
    'J': ('J - Value/Efficiency ($136M)', 136_021_000, 'GRAY'),
}

rosters = {}
for key in roster_files:
    try:
        rosters[key] = pd.read_csv(f'optimized_roster_syn_{key}.csv')
    except FileNotFoundError:
        print(f"  WARNING: optimized_roster_syn_{key}.csv not found, skipping")

print(f"Loaded {len(df_all)} total player-seasons")
print(f"Loaded {len(rosters)} optimized rosters: {list(rosters.keys())}")

# ============================================================
# COLOR PALETTE
# ============================================================

NAVY       = "1B2A4A"
GOLD       = "C9A84C"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
DARK_GRAY  = "2C2C2C"
LIGHT_GRAY = "F2F2F2"
GREEN_COL  = "27AE60"
RED_COL    = "E74C3C"

SCENARIO_COLORS = {
    'A': "1B2A4A",  # Navy
    'B': "1A5276",  # Dark blue
    'C': "1E8449",  # Green
    'D': "6C3483",  # Purple
    'E': "117A65",  # Teal
    'F': "922B21",  # Red
    'G': "935116",  # Brown/orange
    'H': "1A6B7C",  # Cyan-ish
    'I': "4A235A",  # Dark purple
    'J': "424949",  # Gray
}

ARCHETYPE_COLORS = {
    'Elite Playmaker':     'FFD700',  # Gold — franchise cornerstones
    'Two-Way Big':         'D5E8D4',  # Green — versatile two-way bigs
    'Versatile Scorer':    'FFE6CC',  # Orange — versatile guards/wings/bigs
    'Defensive Wing':      'F8CECC',  # Pink-red — defensive wings
    'Perimeter Scorer':    'FFF2CC',  # Yellow — three-point scorers
    'Bench / Role Player': 'F5F5F5',  # Gray — bench/role
}

# ============================================================
# STYLE HELPERS
# ============================================================

def header_cell(ws, row, col, value, bg=NAVY, fg=WHITE, bold=True, size=11):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    return cell

def data_cell(ws, row, col, value, bg=WHITE, bold=False, align='center', fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    if fmt:
        cell.number_format = fmt
    return cell

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

wb = Workbook()

# ============================================================
# TAB 1 — COVER PAGE
# ============================================================

ws_cover = wb.active
ws_cover.title = "Cover"
ws_cover.sheet_view.showGridLines = False

ws_cover.row_dimensions[2].height = 60
ws_cover.row_dimensions[3].height = 40

ws_cover.merge_cells('A2:H2')
title = ws_cover['A2']
title.value = "NBA ROSTER OPTIMIZATION ENGINE"
title.font = Font(bold=True, size=28, color=GOLD)
title.fill = PatternFill("solid", fgColor=NAVY)
title.alignment = Alignment(horizontal='center', vertical='center')

ws_cover.merge_cells('A3:H3')
sub = ws_cover['A3']
sub.value = "Building the Statistically Optimal Team Under a Salary Cap"
sub.font = Font(bold=True, size=16, color=WHITE)
sub.fill = PatternFill("solid", fgColor=NAVY)
sub.alignment = Alignment(horizontal='center', vertical='center')

ws_cover.merge_cells('A5:H5')
ws_cover['A5'].value = "Project Overview"
ws_cover['A5'].font = Font(bold=True, size=13, color=NAVY)
ws_cover['A5'].alignment = Alignment(horizontal='left')

rows_overview = [
    ("Data",        f"{len(df_all):,} player-seasons: 3 seasons x ~400 players (2021-22 to 2023-24). Minimum 20 GP and 10 MPG filter applied."),
    ("Method",      "3-framework Pearson + Spearman correlation -> PCA composite score (0-100) -> K-Means clustering (6 archetypes) -> MILP optimization + Synergy Layer"),
    ("Key Stats",   "9 validated stats: OFF/DEF rating (team-adjusted), ON/OFF Split Diff, Influence Score (USG% x TS%), BLK, STL, AST%, TS%, FG3%"),
    ("Archetypes",  "6 archetypes: Elite Playmaker, Two-Way Big, Versatile Scorer, Defensive Wing, Perimeter Scorer, Bench / Role Player"),
    ("Scenarios",   "10 optimized 15-man rosters: Hard Cap, Luxury Tax, Budget, Rebuild, Win-Now, Defensive Identity, Offensive Identity, Three-Point Era, Small Ball, Value/Efficiency"),
]

for i, (label, desc) in enumerate(rows_overview):
    r = 6 + i
    ws_cover.row_dimensions[r].height = 25
    ws_cover.merge_cells(f'A{r}:B{r}')
    ws_cover.merge_cells(f'C{r}:H{r}')
    ws_cover[f'A{r}'].value = label
    ws_cover[f'A{r}'].font = Font(bold=True, size=10, color=NAVY)
    ws_cover[f'C{r}'].value = desc
    ws_cover[f'C{r}'].font = Font(size=10)
    ws_cover[f'C{r}'].alignment = Alignment(wrap_text=True)

for col in range(1, 9):
    set_col_width(ws_cover, col, 18)

print("Tab 1 (Cover) done")

# ============================================================
# TAB 2 — MASTER PLAYER DATABASE (2023-24) with RANK column
# ============================================================

ws_db = wb.create_sheet("Player Database")
ws_db.sheet_view.showGridLines = False

cols_db = ['PLAYER_NAME', 'TEAM_ABBREVIATION', 'ARCHETYPE',
           'COMPOSITE_SCORE_NORM', 'SALARY', 'VORPD',
           'OFF_RATING_ADJUSTED', 'DEF_RATING_ADJUSTED',
           'ON_OFF_DIFF', 'INFLUENCE_SCORE',
           'TS_PCT', 'BLK', 'STL', 'AST_PCT_ADJUSTED', 'FG3_PCT']

col_labels = ['Rank', 'Player', 'Team', 'Archetype', 'Score (0-100)',
              'Salary', 'VORPD', 'OFF Rtg Adj',
              'DEF Rtg Adj', 'ON/OFF Diff', 'Influence',
              'TS%', 'BLK', 'STL', 'AST% Adj', 'FG3%']

# Title row
ws_db.merge_cells(f'A1:{get_column_letter(len(col_labels)+1)}1')
header_cell(ws_db, 1, 1, f"NBA Player Database - 2023-24 Season ({len(df_2324)} Players, ranked by Composite Score)", bg=NAVY, size=13)
ws_db.row_dimensions[1].height = 30

# Column headers
for c, label in enumerate(col_labels, 1):
    header_cell(ws_db, 2, c, label, bg=DARK_GRAY)
ws_db.row_dimensions[2].height = 35

# Data — sorted by composite score, RANK column added
df_db = df_2324[cols_db].sort_values('COMPOSITE_SCORE_NORM', ascending=False).reset_index(drop=True)

for r_idx, row in df_db.iterrows():
    row_num = r_idx + 3     # row 1=title, row 2=headers, row 3=player #1
    rank    = r_idx + 1     # rank starts at 1
    bg = ARCHETYPE_COLORS.get(row['ARCHETYPE'], WHITE)

    # Rank column
    data_cell(ws_db, row_num, 1, rank, bg=bg, bold=True)

    # Rest of the data columns
    for c_idx, col in enumerate(cols_db, 2):   # starts at column 2 now
        val = row[col]
        if col == 'SALARY':
            data_cell(ws_db, row_num, c_idx, val, bg=bg, fmt='$#,##0')
        elif col in ['COMPOSITE_SCORE_NORM', 'VORPD']:
            data_cell(ws_db, row_num, c_idx, round(float(val), 2) if pd.notna(val) else '', bg=bg)
        elif col in ['TS_PCT', 'FG3_PCT']:
            data_cell(ws_db, row_num, c_idx, round(float(val), 3) if pd.notna(val) else '', bg=bg, fmt='0.000')
        elif col in ['OFF_RATING_ADJUSTED', 'DEF_RATING_ADJUSTED',
                     'AST_PCT_ADJUSTED', 'DREB_PCT_ADJUSTED', 'INFLUENCE_SCORE',
                     'ON_OFF_DIFF', 'BLK', 'STL']:
            data_cell(ws_db, row_num, c_idx, round(float(val), 2) if pd.notna(val) else '', bg=bg, fmt='0.00')
        else:
            data_cell(ws_db, row_num, c_idx, val, bg=bg, align='left' if c_idx <= 4 else 'center')

# Column widths: Rank | Player | Team | Archetype | Score | Salary | VORPD | OFF | DEF | ON/OFF | Influence | TS% | BLK | STL | AST% | FG3%
widths_db = [6, 22, 7, 22, 13, 14, 8, 11, 11, 10, 9, 7, 6, 6, 9, 7]
for c, w in enumerate(widths_db, 1):
    set_col_width(ws_db, c, w)

n_cols_db = len(col_labels) + 1  # +1 for Rank col
ws_db.freeze_panes = ws_db.cell(row=3, column=1)
ws_db.auto_filter.ref = f"A2:{get_column_letter(n_cols_db)}{len(df_db)+2}"
print("Tab 2 (Player Database) done")

# ============================================================
# TAB 2B — ALL SEASONS (2021-22, 2022-23, 2023-24)
# ============================================================

ws_all = wb.create_sheet("All Seasons")
ws_all.sheet_view.showGridLines = False

cols_all = ['SEASON', 'PLAYER_NAME', 'TEAM_ABBREVIATION', 'ARCHETYPE',
            'COMPOSITE_SCORE_NORM', 'SALARY', 'VORPD',
            'OFF_RATING_ADJUSTED', 'DEF_RATING_ADJUSTED',
            'ON_OFF_DIFF', 'INFLUENCE_SCORE', 'TS_PCT', 'BLK', 'STL', 'FG3_PCT']

col_labels_all = ['Season', 'Player', 'Team', 'Archetype', 'Score (0-100)',
                  'Salary (est.)', 'VORPD', 'OFF Rtg Adj', 'DEF Rtg Adj',
                  'ON/OFF Diff', 'Influence', 'TS%', 'BLK', 'STL', 'FG3%']

ws_all.merge_cells(f'A1:{get_column_letter(len(col_labels_all))}1')
header_cell(ws_all, 1, 1,
            f"All Seasons Player Database - {len(df_all):,} Player-Seasons (2021-22, 2022-23, 2023-24)",
            bg=NAVY, size=13)
ws_all.row_dimensions[1].height = 30

for c, label in enumerate(col_labels_all, 1):
    header_cell(ws_all, 2, c, label, bg=DARK_GRAY)
ws_all.row_dimensions[2].height = 35

# Sort: season desc, then score desc
df_all_display = (df_all[cols_all]
                  .sort_values(['SEASON', 'COMPOSITE_SCORE_NORM'],
                               ascending=[False, False])
                  .reset_index(drop=True))

# Season background colors so seasons are visually separated
season_bg = {
    '2023-24': None,   # use archetype color
    '2022-23': 'EBF5FB',  # light blue tint for 2022-23
    '2021-22': 'EAFAF1',  # light green tint for 2021-22
}

for r_idx, row in df_all_display.iterrows():
    row_num = r_idx + 3
    arch_bg  = ARCHETYPE_COLORS.get(row['ARCHETYPE'], WHITE)
    season_tint = season_bg.get(row['SEASON'])
    bg = arch_bg if season_tint is None else season_tint

    for c_idx, col in enumerate(cols_all, 1):
        val = row[col]
        if col == 'SALARY':
            data_cell(ws_all, row_num, c_idx, val, bg=bg, fmt='$#,##0')
        elif col in ['COMPOSITE_SCORE_NORM', 'VORPD']:
            data_cell(ws_all, row_num, c_idx,
                      round(float(val), 2) if pd.notna(val) else '', bg=bg)
        elif col in ['TS_PCT', 'FG3_PCT']:
            data_cell(ws_all, row_num, c_idx,
                      round(float(val), 3) if pd.notna(val) else '', bg=bg, fmt='0.000')
        elif col in ['OFF_RATING_ADJUSTED', 'DEF_RATING_ADJUSTED',
                     'ON_OFF_DIFF', 'INFLUENCE_SCORE', 'BLK', 'STL']:
            data_cell(ws_all, row_num, c_idx,
                      round(float(val), 2) if pd.notna(val) else '', bg=bg, fmt='0.00')
        elif col == 'SEASON':
            data_cell(ws_all, row_num, c_idx, val, bg=bg, bold=True)
        else:
            data_cell(ws_all, row_num, c_idx, val, bg=bg,
                      align='left' if c_idx <= 4 else 'center')

widths_all = [9, 22, 7, 22, 13, 14, 8, 11, 11, 10, 9, 7, 6, 6, 7]
for c, w in enumerate(widths_all, 1):
    set_col_width(ws_all, c, w)

ws_all.freeze_panes = ws_all.cell(row=3, column=1)
ws_all.auto_filter.ref = f"A2:{get_column_letter(len(cols_all))}{len(df_all_display)+2}"
print("Tab 3 (All Seasons) done")

# ============================================================
# TAB 3 (now TAB 4) — OPTIMIZED ROSTERS (ALL 10 SCENARIOS)
# Each scenario gets its own mini-table stacked vertically
# ============================================================

ws_ros = wb.create_sheet("Optimized Rosters")
ws_ros.sheet_view.showGridLines = False

ws_ros.merge_cells('A1:J1')
header_cell(ws_ros, 1, 1, "Optimized Rosters — Synergy Model (Syn Score = Base Score + NET_SYNERGY_PROFILE adjustment)", bg=NAVY, size=12)
ws_ros.row_dimensions[1].height = 28

current_row = 2

for key, roster in rosters.items():
    label, cap, _ = roster_files[key]
    color = SCENARIO_COLORS[key]

    # Scenario title bar
    ws_ros.merge_cells(f'A{current_row}:J{current_row}')
    header_cell(ws_ros, current_row, 1, f"SCENARIO {label}", bg=color, size=12)
    ws_ros.row_dimensions[current_row].height = 28
    current_row += 1

    # Column headers
    headers = ['#', 'Player', 'Team', 'Archetype', 'Syn Score', 'Base Score', 'NET_SYN', 'Salary', 'VORPD', 'Age']
    for c, h in enumerate(headers, 1):
        header_cell(ws_ros, current_row, c, h, bg=DARK_GRAY)
    current_row += 1

    # Player rows — use COMPOSITE_SYNERGY (synergy-adjusted) as primary score
    score_col = 'COMPOSITE_SYNERGY' if 'COMPOSITE_SYNERGY' in roster.columns else 'COMPOSITE_SCORE_NORM'
    sorted_roster = roster.sort_values(score_col, ascending=False).reset_index(drop=True)
    total_salary    = sorted_roster['SALARY'].sum()
    total_syn_score = sorted_roster[score_col].sum()
    total_base_score= sorted_roster['COMPOSITE_SCORE_NORM'].sum()

    for i, row in sorted_roster.iterrows():
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        net_syn = round(row['NET_SYNERGY_PROFILE'], 2) if 'NET_SYNERGY_PROFILE' in row and pd.notna(row['NET_SYNERGY_PROFILE']) else ''
        age_val = row['AGE'] if 'AGE' in row and pd.notna(row['AGE']) else ''
        data_cell(ws_ros, current_row, 1, i + 1, bg=bg, bold=True)
        data_cell(ws_ros, current_row, 2, row['PLAYER_NAME'], bg=bg, align='left')
        data_cell(ws_ros, current_row, 3, row['TEAM_ABBREVIATION'], bg=bg)
        data_cell(ws_ros, current_row, 4, row['ARCHETYPE'], bg=bg, align='left')
        data_cell(ws_ros, current_row, 5, round(row[score_col], 2), bg=bg)
        data_cell(ws_ros, current_row, 6, round(row['COMPOSITE_SCORE_NORM'], 2), bg=bg)
        data_cell(ws_ros, current_row, 7, net_syn, bg=bg, fmt='0.00')
        data_cell(ws_ros, current_row, 8, row['SALARY'], bg=bg, fmt='$#,##0')
        data_cell(ws_ros, current_row, 9, round(row['VORPD'], 2) if pd.notna(row['VORPD']) else '', bg=bg)
        data_cell(ws_ros, current_row, 10, age_val, bg=bg)
        current_row += 1

    # Totals row
    ws_ros.merge_cells(f'A{current_row}:D{current_row}')
    header_cell(ws_ros, current_row, 1, f"TOTALS  |  Cap used: ${total_salary:,.0f} / ${cap:,.0f}  ({total_salary/cap*100:.1f}%)", bg=color, fg=WHITE)
    data_cell(ws_ros, current_row, 5, round(total_syn_score, 2), bg=LIGHT_BLUE, bold=True)
    data_cell(ws_ros, current_row, 6, round(total_base_score, 2), bg=LIGHT_BLUE, bold=True)
    data_cell(ws_ros, current_row, 8, total_salary, bg=LIGHT_BLUE, bold=True, fmt='$#,##0')
    current_row += 2  # blank line between scenarios

widths_ros = [5, 24, 7, 24, 11, 11, 9, 16, 8, 6]
for c, w in enumerate(widths_ros, 1):
    set_col_width(ws_ros, c, w)

ws_ros.freeze_panes = ws_ros.cell(row=3, column=1)
print("Tab 3 (Optimized Rosters) done")

# ============================================================
# TAB 4 — SCENARIO COMPARISON (all 10 side by side)
# ============================================================

ws_cmp = wb.create_sheet("Scenario Comparison")
ws_cmp.sheet_view.showGridLines = False

ws_cmp.merge_cells(f'A1:{get_column_letter(1 + len(rosters))}1')
header_cell(ws_cmp, 1, 1, "Scenario Comparison - All 10 Rosters", bg=NAVY, size=13)
ws_cmp.row_dimensions[1].height = 30

# Header row: Metric | A | B | C | D | E | F | G | H | I | J
data_cell(ws_cmp, 2, 1, "Metric", bg=DARK_GRAY, bold=True, align='left')
for c_idx, key in enumerate(rosters.keys(), 2):
    label = roster_files[key][0]
    header_cell(ws_cmp, 2, c_idx, label, bg=SCENARIO_COLORS[key], size=9)
ws_cmp.row_dimensions[2].height = 40

# Metrics
def get_stats(roster, cap):
    score_col = 'COMPOSITE_SYNERGY' if 'COMPOSITE_SYNERGY' in roster.columns else 'COMPOSITE_SCORE_NORM'
    return {
        'total_score': roster[score_col].sum(),
        'avg_score':   roster[score_col].mean(),
        'max_score':   roster[score_col].max(),
        'total_sal':   roster['SALARY'].sum(),
        'cap_used':    roster['SALARY'].sum() / cap * 100,
        'avg_vorpd':   roster['VORPD'].mean() if 'VORPD' in roster.columns else 0,
        'n_stars':     int((roster[score_col] >= 70).sum()),
        'avg_age':     roster['AGE'].mean() if 'AGE' in roster.columns else 0,
    }

stats = {key: get_stats(rosters[key], roster_files[key][1]) for key in rosters}

metric_rows = [
    ('Total Synergy Score',  'total_score', '0.00'),
    ('Average Synergy Score','avg_score',   '0.00'),
    ('Best Synergy Score',   'max_score',   '0.00'),
    ('Total Salary Used',    'total_sal',   '$#,##0'),
    ('Cap % Used',           'cap_used',    '0.0'),
    ('Average VORPD',        'avg_vorpd',   '0.00'),
    ('Players Scoring 70+',  'n_stars',     '0'),
    ('Average Age',          'avg_age',     '0.0'),
]

for i, (label, metric, fmt) in enumerate(metric_rows):
    r = i + 3
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    data_cell(ws_cmp, r, 1, label, bg=bg, bold=True, align='left')
    for c_idx, key in enumerate(rosters.keys(), 2):
        val = stats[key][metric]
        cell = data_cell(ws_cmp, r, c_idx, round(val, 2) if isinstance(val, float) else val, bg=bg, fmt=fmt)

set_col_width(ws_cmp, 1, 24)
for c in range(2, len(rosters) + 2):
    set_col_width(ws_cmp, c, 18)

print("Tab 4 (Scenario Comparison) done")

# ============================================================
# TAB 5 — ARCHETYPE SUMMARY
# ============================================================

ws_arch = wb.create_sheet("Archetypes")
ws_arch.sheet_view.showGridLines = False

ws_arch.merge_cells('A1:I1')
header_cell(ws_arch, 1, 1, "Player Archetypes - K-Means Clustering Results (6 Archetypes)", bg=NAVY, size=13)
ws_arch.row_dimensions[1].height = 30

arch_headers = ['Archetype', 'Players\n(3 seasons)', 'Players\n(2023-24)',
                'Avg Score', 'Avg Salary', 'Avg VORPD',
                'Avg OFF Rtg', 'Avg DEF Rtg', 'Defining Traits']
for c, h in enumerate(arch_headers, 1):
    header_cell(ws_arch, 2, c, h, bg=DARK_GRAY)
ws_arch.row_dimensions[2].height = 35

arch_traits = {
    'Elite Playmaker':     'Highest OFF_RATING_ADJUSTED + ON_OFF_DIFF. Franchise cornerstones who elevate everyone. Jokic, Embiid, SGA, Giannis, Luka.',
    'Two-Way Big':         'High BLK + balanced OFF/DEF impact. Versatile bigs who anchor both ends. Wembanyama, AD, Gobert, Nurkic, Hartenstein, Chet.',
    'Versatile Scorer':    'High STL + AST%. Positionally diverse scorers and playmakers — guards, wings, and forwards. Franz Wagner, Garland, Randle, Giddey.',
    'Defensive Wing':      'High court impact (ON_OFF_DIFF) on weaker defensive teams. OG Anunoby, KCP, Tari Eason, Aaron Gordon.',
    'Perimeter Scorer':    'High FG3%. Outside scoring and spacing wings. Markkanen, RJ Barrett, Grayson Allen, Kuminga.',
    'Bench / Role Player': 'Below average across all stats. Lowest ON_OFF_DIFF. End-of-bench depth pieces.',
}

for i, (arch, trait) in enumerate(arch_traits.items()):
    r = i + 3
    bg = list(ARCHETYPE_COLORS.values())[i]
    players_all  = df_all[df_all['ARCHETYPE'] == arch]
    players_2324 = df_2324[df_2324['ARCHETYPE'] == arch]

    data_cell(ws_arch, r, 1, arch, bg=bg, bold=True, align='left')
    data_cell(ws_arch, r, 2, len(players_all), bg=bg)
    data_cell(ws_arch, r, 3, len(players_2324), bg=bg)
    data_cell(ws_arch, r, 4, round(players_2324['COMPOSITE_SCORE_NORM'].mean(), 1), bg=bg)
    data_cell(ws_arch, r, 5, int(players_2324['SALARY'].mean()) if players_2324['SALARY'].notna().any() else 0, bg=bg, fmt='$#,##0')
    data_cell(ws_arch, r, 6, round(players_2324['VORPD'].mean(), 2), bg=bg)
    data_cell(ws_arch, r, 7, round(players_2324['OFF_RATING_ADJUSTED'].mean(), 2), bg=bg, fmt='0.00')
    data_cell(ws_arch, r, 8, round(players_2324['DEF_RATING_ADJUSTED'].mean(), 2), bg=bg, fmt='0.00')
    data_cell(ws_arch, r, 9, trait, bg=bg, align='left')

for c, w in enumerate([24, 10, 10, 10, 14, 10, 11, 11, 45], 1):
    set_col_width(ws_arch, c, w)

print("Tab 5 (Archetypes) done")

# ============================================================
# TAB 6 — METHODOLOGY
# ============================================================

ws_meth = wb.create_sheet("Methodology")
ws_meth.sheet_view.showGridLines = False

ws_meth.merge_cells('A1:F1')
header_cell(ws_meth, 1, 1, "Methodology - NBA Roster Optimization Engine", bg=NAVY, size=14)
ws_meth.row_dimensions[1].height = 35

methodology_text = [
    ("DATA COLLECTION", ""),
    ("Source", "NBA Stats API (nba_api Python package) - official NBA.com backend"),
    ("Seasons", f"2021-22, 2022-23, 2023-24 (3 seasons x ~400 players = {len(df_all):,} player-seasons after quality filter)"),
    ("Filter", "Minimum 20 games played and 10 minutes per game to exclude noise"),
    ("Salary", "Basketball Reference contracts scaled to historical seasons via cap-ratio method"),
    ("Recency Weights", "2021-22: 20%, 2022-23: 35%, 2023-24: 45% - recent seasons weighted more"),
    ("", ""),
    ("FEATURE ENGINEERING", ""),
    ("Team Adjustment", "All on-court ratings adjusted by subtracting team average to isolate individual contribution from teammate quality"),
    ("TS% Filter", "Players with fewer than 100 total field goal attempts receive league average TS% to prevent small-sample distortion"),
    ("INFLUENCE_SCORE", "USG% x TS% = productive offensive load. Rewards high-usage players who also convert efficiently. Separates volume scorers (LaMelo) from efficient high-usage players (Curry, SGA)."),
    ("ON/OFF Split", "5-man lineup data (LeagueDashLineups API). ON = weighted avg team NET_RATING in lineups containing the player. OFF = without. ON_OFF_DIFF = ON - OFF. Captures gravity, spacing, and court impact beyond box scores."),
    ("", ""),
    ("CORRELATION ANALYSIS (3 FRAMEWORKS)", ""),
    ("Framework 1", "Pearson: team-adjusted stats vs W_PCT - identifies stats that predict team winning"),
    ("Framework 2", "Pearson: individual counting stats vs W_PCT - individual skills tied to winning"),
    ("Framework 3", "Pearson: stats vs PLUS_MINUS_ADJUSTED - captures invisible contributions"),
    ("Spearman Validation", "All candidate stats re-validated using Spearman rank correlation to catch non-linear relationships"),
    ("Selection Rule", "Stat included if p < 0.05 in at least one framework AND validated by Spearman"),
    ("", ""),
    ("PCA COMPOSITE SCORE", ""),
    ("Standardization", "All 9 final stats normalized to mean=0, std=1 via StandardScaler before PCA"),
    ("PCA Method", "Multi-component PCA on 9 validated stats. PC1+PC2+PC3 used (65% variance explained vs 30.5% for PC1 alone) — each component weighted by its share of variance explained (PC1: ~47%, PC2: ~29%, PC3: ~24%). This ensures PC1's dominant basketball excellence signal (ON_OFF_DIFF, OFF_RATING, INFLUENCE_SCORE) carries the most weight while PC2/PC3 add independent information. PC1 weights: OFF_RATING_ADJ (0.497), ON_OFF_DIFF (0.435), INFLUENCE_SCORE (0.389), STL (0.387), TS% (0.292). DREB_PCT excluded: position-group adjustment revealed it fails Spearman (rho=0.025, p=0.365) — positional height was the confound, not rebounding skill."),
    ("Score Normalization", "Raw PCA scores rescaled to 0-100 for interpretability"),
    ("VORPD", "Value Over Replacement Per Dollar = (Score - 10th percentile) / (Salary in $M). Measures salary efficiency."),
    ("", ""),
    ("K-MEANS CLUSTERING", ""),
    ("Method", f"K-Means clustering (K=7) on 9 standardized stats across {len(df_all):,} player-seasons. 7 clusters mapped to 6 archetype labels (2 clusters merge into Bench / Role Player)."),
    ("K Selection", "Elbow method used to validate K=6 as the point of diminishing returns"),
    ("Archetypes", "6 clusters manually labeled based on average stat profiles and basketball identity of members"),
    ("", ""),
    ("MILP OPTIMIZATION (10 SCENARIOS)", ""),
    ("Model Type", "Mixed Integer Linear Programming (MILP) via PuLP/CBC solver"),
    ("Decision Variables", "399 binary x_i variables (player selection) + ~1,330 binary y_ij variables (pairwise synergy linearization)"),
    ("Base Constraints", "Salary cap, roster size (15), archetype minimums, max 2 players per NBA team, max 5 big men"),
    ("Scenario A", "Composite + synergy objective | Hard cap: $136M"),
    ("Scenario B", "Composite + synergy objective | Luxury tax: $165M"),
    ("Scenario C", "Composite + synergy objective | Budget: $90M"),
    ("Scenario D", "Composite + synergy | Hard cap + min 9 players age <= 24 (Rebuild)"),
    ("Scenario E", "Composite + synergy | Luxury cap + min 9 players age >= 28 (Win-Now)"),
    ("Scenario F", "Maximize defensive rating | Hard cap (Defensive Identity)"),
    ("Scenario G", "Maximize offensive rating | Hard cap (Offensive Identity)"),
    ("Scenario H", "Composite + synergy | Hard cap + min 10 above-avg 3PT shooters (Three-Point Era)"),
    ("Scenario I", "Composite + synergy | Hard cap + max 1 Two-Way Big (True Small Ball)"),
    ("Scenario J", "Maximize VORPD (value/dollar) | Hard cap (Moneyball)"),
    ("", ""),
    ("SYNERGY MODEL", ""),
    ("Overview", "Three-layer synergy model built on top of individual composite scores. All synergy terms use W_NET_SYNERGY — chosen because it has the highest W_PCT correlation (r=+0.568) across all 3 seasons in validation testing. Offensive and defensive synergy are treated with equal weight."),
    ("Layer 1: Player Profile", "Each player's NET_SYNERGY_PROFILE = their average NET_SYNERGY across all observed 2-man pairs (2023-24). Added to composite score with weight 0.4. Captures whether a player consistently elevates or drags teammates on both ends."),
    ("Layer 2: Pairwise Bonus", "For each pair (i,j) both selected, the optimizer earns W_NET_SYNERGY_SCALED bonus in the objective. MILP linearized via binary y_ij variables: y_ij ≤ x_i, y_ij ≤ x_j, y_ij ≥ x_i+x_j-1. Only pairs with |NET_SYNERGY| >= 1.0 pts/100 poss included to limit variable count."),
    ("Layer 3: Coverage Zones", "Hard defensive coverage constraints: Paint (≥1 player BLK≥1.5), Perimeter (≥2 players STL≥1.3), Switchability (≥4 players from defensive archetypes: Two-Way Big, Versatile Scorer, Defensive Wing)."),
    ("Data Source", "2-man lineup data from NBA Stats API (leaguedashlineups, group_quantity=2). 6,000 pair-seasons across 3 seasons, filtered to MIN≥100 shared minutes for synergy computation."),
    ("Cross-Team Limitation", "Pairs who never played together receive no pairwise bonus — only the player-level profile adjustment. The pairwise bonus is therefore conservative for cross-team rosters (all 10 scenarios are cross-team constructions)."),
    ("", ""),
    ("LIMITATIONS", ""),
    ("Salary Approximation", "Historical salaries estimated via cap-ratio scaling. 87 salary imputation errors corrected (2023-24) using verified contract data — including high-profile errors like Jimmy Butler ($45.2M), Malcolm Brogdon ($22.6M), and Bojan Bogdanovic ($18M). ~23 genuine league-minimum players remain at $1,119,563."),
    ("Defensive Measurement", "Individual defensive stats are difficult to isolate from team context. DEF_RATING is shared across all 5 players simultaneously — the model cannot assign individual credit. PCA weight of 0.039 for DEF_RATING_ADJUSTED reflects this honest limitation, not a modeling flaw. STL (0.366) and BLK (0.274) carry the defensive signal because they are individually attributable."),
    ("Advanced Defensive Stats", "Advanced individual defensive metrics (D-EPM, D-LEBRON) require proprietary Second Spectrum tracking data available only through paid subscriptions. Public APIs (NBA Stats, Basketball Reference) do not expose this data. This is an acknowledged constraint of using public data only."),
    ("Team Context Limitation", "PLUS_MINUS still contains lineup contamination bias that team-averaging cannot fully remove."),
    ("Player Selection Scope", "Composite scores are built from all 3 seasons (1,185 player-seasons) with recency weights (2021-22: 20%, 2022-23: 35%, 2023-24: 45%). The optimizer selects from 2023-24 players only — it reflects current roster construction, not multi-year projections. Player aging curves not modeled."),
]

for i, (label, text) in enumerate(methodology_text):
    r = i + 2
    if text == "" and label == "":
        ws_meth.row_dimensions[r].height = 8
        continue
    if text == "":
        ws_meth.merge_cells(f'A{r}:F{r}')
        cell = ws_meth[f'A{r}']
        cell.value = label
        cell.font = Font(bold=True, size=11, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK_GRAY)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws_meth.row_dimensions[r].height = 22
    else:
        ws_meth[f'A{r}'].value = label
        ws_meth[f'A{r}'].font = Font(bold=True, size=10, color=NAVY)
        ws_meth[f'A{r}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws_meth.merge_cells(f'B{r}:F{r}')
        ws_meth[f'B{r}'].value = text
        ws_meth[f'B{r}'].font = Font(size=10)
        ws_meth[f'B{r}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        ws_meth[f'A{r}'].fill = PatternFill("solid", fgColor=bg)
        ws_meth[f'B{r}'].fill = PatternFill("solid", fgColor=bg)
        ws_meth.row_dimensions[r].height = 30

set_col_width(ws_meth, 1, 28)
for c in range(2, 7):
    set_col_width(ws_meth, c, 20)

print("Tab 6 (Methodology) done")

# ============================================================
# TAB 7 — ANALYTICAL CHARTS
# Embeds all pipeline-generated charts into the dashboard
# so everything lives in one file for presentation.
# ============================================================

from openpyxl.drawing.image import Image as XlImage
import os

ws_charts = wb.create_sheet("Charts")
ws_charts.sheet_view.showGridLines = False

ws_charts.merge_cells('A1:N1')
header_cell(ws_charts, 1, 1, "Analytical Charts — Statistical Process Visualization", bg=NAVY, size=14)
ws_charts.row_dimensions[1].height = 35

chart_files = [
    ('charts/top20_players_final.png',       'Top 20 Players by Composite Score (2023-24)',
     'The final ranking output — PCA-derived composite scores on a 0-100 scale. Jokic, SGA, and Luka lead.'),
    ('charts/pca_weights_final.png',         'PCA-Derived Stat Weights',
     'How much each stat contributes to the composite score. Weights are entirely data-driven — OFF_RATING_ADJUSTED and ON_OFF_DIFF dominate.'),
    ('charts/pca_variance_explained.png',    'PCA Variance Explained Per Component',
     'PC1 captures ~30% of variance alone. We use PC1+PC2+PC3 (65% combined) weighted by variance share to capture more basketball signal.'),
    ('charts/intercorrelation_final.png',    'Intercorrelation Matrix — Final 9 Stats',
     'Checks that no two stats are redundant (|r| > 0.7). All 9 stats passed — each captures an independent dimension of player skill.'),
    ('charts/elbow_method.png',              'Elbow Method — Optimal Number of Clusters',
     'Inertia drops steeply until K=7, then flattens. K=7 balances granularity with interpretability (7 clusters → 6 archetype names).'),
    ('charts/cluster_visualization.png',     'Player Archetypes — K-Means Clustering (PCA 2D Projection)',
     'All 1,185 player-seasons projected onto 2 PCA dimensions. Colors = archetypes. Notable players annotated. Shows natural groupings.'),
    ('charts/archetype_radar_charts.png',    'Archetype Stat Profiles — Radar Charts',
     'Average stat profile per archetype on a 0-1 normalized scale. Each archetype has a distinct fingerprint — Elite Playmakers excel at offense, Two-Way Bigs at blocks, etc.'),
]

current_row = 3

for filepath, title, description in chart_files:
    if not os.path.exists(filepath):
        print(f"  WARNING: {filepath} not found, skipping")
        continue

    # Chart title
    ws_charts.merge_cells(f'A{current_row}:N{current_row}')
    cell = ws_charts[f'A{current_row}']
    cell.value = title
    cell.font = Font(bold=True, size=12, color=NAVY)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws_charts.row_dimensions[current_row].height = 25
    current_row += 1

    # Chart description
    ws_charts.merge_cells(f'A{current_row}:N{current_row}')
    cell = ws_charts[f'A{current_row}']
    cell.value = description
    cell.font = Font(size=10, italic=True, color='555555')
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws_charts.row_dimensions[current_row].height = 30
    current_row += 1

    # Insert image
    img = XlImage(filepath)
    img.width = 720
    img.height = 430
    ws_charts.add_image(img, f'A{current_row}')
    # Reserve ~22 rows for each chart image
    for r in range(current_row, current_row + 22):
        ws_charts.row_dimensions[r].height = 20
    current_row += 24  # 22 for image + 2 spacing

for c in range(1, 15):
    set_col_width(ws_charts, c, 10)

print("Tab 7 (Charts) done")

# ============================================================
# SAVE
# ============================================================

output_file = 'NBA_Optimizer_Dashboard.xlsx'
wb.save(output_file)

print(f"\n{'=' * 60}")
print(f"DASHBOARD SAVED: {output_file}")
print(f"{'=' * 60}")
print("\nTabs:")
print("  1. Cover               - Project overview (10 scenarios)")
print("  2. Player Database     - 399 players ranked #1-399 by composite score")
print("  3. All Seasons         - All 1,185 player-seasons across 3 years")
print("  4. Optimized Rosters   - All 10 salary cap scenarios")
print("  5. Scenario Comparison - Side-by-side metrics for all 10")
print("  6. Archetypes          - K-Means cluster summary")
print("  7. Methodology         - Full statistical methodology")
print("  8. Charts              - All 7 analytical charts with descriptions")
